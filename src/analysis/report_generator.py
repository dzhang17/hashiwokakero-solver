"""
Report generator for benchmark results.
"""

import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Any, Optional
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference, Series
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import seaborn as sns
from io import BytesIO

from ..visualization.performance_viz import PerformanceVisualizer
from ..core.utils import setup_logger


class ReportGenerator:
    """Generate comprehensive Excel reports from benchmark results"""
    
    def __init__(self, results_file: Path, output_dir: Optional[Path] = None):
        """
        Initialize report generator.
        
        Args:
            results_file: Path to benchmark results CSV/JSON
            output_dir: Output directory for reports
        """
        self.logger = setup_logger(self.__class__.__name__)
        
        # Load results
        if results_file.suffix == '.csv':
            self.results_df = pd.read_csv(results_file)
        else:
            with open(results_file, 'r') as f:
                data = json.load(f)
                self.results_df = pd.DataFrame(data['results'])
                self.config = data.get('config', {})
                self.summary = data.get('summary', {})
                
        self.output_dir = output_dir or Path('results/reports')
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
        # Visualizer for creating charts
        self.viz = PerformanceVisualizer()
        
    def generate_excel_report(self, filename: Optional[str] = None) -> Path:
        """
        Generate comprehensive Excel report.
        
        Args:
            filename: Optional filename for the report
            
        Returns:
            Path to generated report
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"benchmark_report_{timestamp}.xlsx"
            
        filepath = self.output_dir / filename
        
        # Create workbook
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create sheets
        self._create_summary_sheet(wb)
        self._create_detailed_results_sheet(wb)
        self._create_algorithm_comparison_sheet(wb)
        self._create_performance_charts_sheet(wb)
        self._create_statistics_sheet(wb)
        
        # Save workbook
        wb.save(filepath)
        self.logger.info(f"Excel report saved to {filepath}")
        
        return filepath
        
    def _create_summary_sheet(self, wb: Workbook):
        """Create summary sheet with key metrics"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "Hashiwokakero Solver Benchmark Report"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:F1')
        
        # Report metadata
        ws['A3'] = "Report Generated:"
        ws['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A4'] = "Total Tests:"
        ws['B4'] = len(self.results_df)
        ws['A5'] = "Algorithms Tested:"
        ws['B5'] = ", ".join(self.results_df['algorithm'].unique())
        
        # Overall statistics
        ws['A7'] = "Overall Statistics"
        ws['A7'].font = Font(size=14, bold=True)
        
        row = 9
        ws.cell(row=row, column=1, value="Metric")
        ws.cell(row=row, column=2, value="Value")
        for cell in ws[row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
        row += 1
        
        # Calculate and add overall metrics
        metrics = {
            "Overall Success Rate": f"{self.results_df['success'].mean() * 100:.1f}%",
            "Average Solving Time": f"{self.results_df['solve_time'].mean():.2f}s",
            "Median Solving Time": f"{self.results_df['solve_time'].median():.2f}s",
            "Average Memory Usage": f"{self.results_df['memory_mb'].mean():.1f} MB",
            "Valid Solutions": f"{self.results_df['is_valid'].sum()} / {self.results_df['success'].sum()}",
        }
        
        for metric, value in metrics.items():
            ws.cell(row=row, column=1, value=metric)
            ws.cell(row=row, column=2, value=value)
            row += 1
            
        # Algorithm comparison summary
        row += 2
        ws.cell(row=row, column=1, value="Algorithm Performance Summary")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        
        row += 2
        summary_df = self.results_df.groupby('algorithm').agg({
            'success': lambda x: f"{x.mean() * 100:.1f}%",
            'solve_time': lambda x: f"{x.mean():.2f}s",
            'memory_mb': lambda x: f"{x.mean():.1f} MB"
        }).reset_index()
        
        summary_df.columns = ['Algorithm', 'Success Rate', 'Avg Time', 'Avg Memory']
        
        # Write summary table
        for r in dataframe_to_rows(summary_df, index=False, header=True):
            ws.append(r)
            
        # Format the summary table
        for cell in ws[row + 2]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def _create_detailed_results_sheet(self, wb: Workbook):
        """Create sheet with detailed test results"""
        ws = wb.create_sheet("Detailed Results")
        
        # Prepare data
        detailed_df = self.results_df.copy()
        
        # Select and order columns
        columns = ['puzzle_id', 'algorithm', 'width', 'height', 'num_islands', 
                  'difficulty', 'success', 'solve_time', 'iterations', 
                  'memory_mb', 'is_valid']
        
        detailed_df = detailed_df[columns]
        
        # Write data
        for r in dataframe_to_rows(detailed_df, index=False, header=True):
            ws.append(r)
            
        # Format header
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(color="FFFFFF", bold=True)
            
        # Add filters
        ws.auto_filter.ref = ws.dimensions
        
        # Conditional formatting for success column
        from openpyxl.formatting.rule import CellIsRule
        
        # Find success column
        for idx, cell in enumerate(ws[1]):
            if cell.value == 'success':
                col_letter = cell.column_letter
                ws.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}{ws.max_row}',
                    CellIsRule(operator='equal', formula=['TRUE'], 
                             fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
                )
                ws.conditional_formatting.add(
                    f'{col_letter}2:{col_letter}{ws.max_row}',
                    CellIsRule(operator='equal', formula=['FALSE'], 
                             fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
                )
                
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def _create_algorithm_comparison_sheet(self, wb: Workbook):
        """Create algorithm comparison sheet"""
        ws = wb.create_sheet("Algorithm Comparison")
        
        # Title
        ws['A1'] = "Algorithm Comparison Analysis"
        ws['A1'].font = Font(size=16, bold=True)
        
        # Comparison by difficulty
        row = 3
        ws.cell(row=row, column=1, value="Performance by Difficulty")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        
        row += 2
        diff_pivot = pd.pivot_table(
            self.results_df,
            values=['success', 'solve_time'],
            index='difficulty',
            columns='algorithm',
            aggfunc={'success': 'mean', 'solve_time': 'mean'}
        )
        
        # Write difficulty comparison
        ws.cell(row=row, column=1, value="Success Rate by Difficulty")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        # Convert success rates to percentages
        success_df = (diff_pivot['success'] * 100).round(1)
        success_df = success_df.reset_index()
        
        for r in dataframe_to_rows(success_df, index=False, header=True):
            ws.append(r)
            current_row = ws.max_row
            for cell in ws[current_row]:
                if isinstance(cell.value, (int, float)):
                    cell.value = f"{cell.value}%"
                    
        # Time comparison
        row = ws.max_row + 3
        ws.cell(row=row, column=1, value="Average Solving Time by Difficulty (seconds)")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        time_df = diff_pivot['solve_time'].round(2).reset_index()
        
        for r in dataframe_to_rows(time_df, index=False, header=True):
            ws.append(r)
            
        # Comparison by size
        row = ws.max_row + 3
        ws.cell(row=row, column=1, value="Performance by Problem Size")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        
        row += 2
        size_comparison = self.results_df.groupby(['num_islands', 'algorithm']).agg({
            'success': 'mean',
            'solve_time': 'mean',
            'memory_mb': 'mean'
        }).round(2).reset_index()
        
        # Pivot for better presentation
        size_pivot = size_comparison.pivot(
            index='num_islands',
            columns='algorithm',
            values=['success', 'solve_time', 'memory_mb']
        )
        
        # Write size comparison tables
        metrics = ['success', 'solve_time', 'memory_mb']
        metric_names = ['Success Rate', 'Avg Time (s)', 'Avg Memory (MB)']
        
        for metric, name in zip(metrics, metric_names):
            ws.cell(row=row, column=1, value=f"{name} by Problem Size")
            ws.cell(row=row, column=1).font = Font(bold=True)
            row += 1
            
            metric_df = size_pivot[metric].reset_index()
            
            for r in dataframe_to_rows(metric_df, index=False, header=True):
                ws.append(r)
                if metric == 'success':
                    current_row = ws.max_row
                    for cell in ws[current_row][1:]:  # Skip first column
                        if isinstance(cell.value, (int, float)):
                            cell.value = f"{cell.value * 100:.1f}%"
                            
            row = ws.max_row + 2
            
    def _create_performance_charts_sheet(self, wb: Workbook):
        """Create sheet with performance charts"""
        ws = wb.create_sheet("Performance Charts")
        
        # Create charts using matplotlib
        charts_to_create = [
            ('time_comparison', 'Solving Time by Algorithm'),
            ('success_rate', 'Success Rate by Algorithm and Difficulty'),
            ('scalability', 'Scalability Analysis'),
            ('memory_usage', 'Memory Usage Comparison')
        ]
        
        row = 1
        for chart_type, title in charts_to_create:
            # Create chart
            fig = self._create_chart(chart_type)
            
            if fig:
                # Save to BytesIO
                img_buffer = BytesIO()
                fig.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
                img_buffer.seek(0)
                plt.close(fig)
                
                # Add to worksheet
                img = Image(img_buffer)
                img.width = 600
                img.height = 400
                ws.add_image(img, f'A{row}')
                
                row += 22  # Space for next chart
                
    def _create_chart(self, chart_type: str) -> Optional[plt.Figure]:
        """Create specific chart type"""
        if chart_type == 'time_comparison':
            return self.viz.plot_time_comparison(self.results_df)
            
        elif chart_type == 'success_rate':
            return self.viz.plot_success_rate(self.results_df)
            
        elif chart_type == 'scalability':
            return self.viz.plot_scalability(self.results_df)
            
        elif chart_type == 'memory_usage':
            return self.viz.plot_memory_usage(self.results_df)
            
        return None
        
    def _create_statistics_sheet(self, wb: Workbook):
        """Create detailed statistics sheet"""
        ws = wb.create_sheet("Statistics")
        
        # Title
        ws['A1'] = "Detailed Statistical Analysis"
        ws['A1'].font = Font(size=16, bold=True)
        
        row = 3
        
        # Algorithm statistics
        ws.cell(row=row, column=1, value="Algorithm Statistics")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        
        row += 2
        
        # Detailed statistics per algorithm
        stats_df = self.results_df.groupby('algorithm').agg({
            'success': ['count', 'sum', 'mean'],
            'solve_time': ['mean', 'median', 'std', 'min', 'max'],
            'memory_mb': ['mean', 'median', 'max'],
            'iterations': ['mean', 'median', 'max']
        }).round(3)
        
        # Flatten column names
        stats_df.columns = ['_'.join(col).strip() for col in stats_df.columns.values]
        stats_df = stats_df.reset_index()
        
        # Rename columns for clarity
        column_mapping = {
            'success_count': 'Total Tests',
            'success_sum': 'Successful',
            'success_mean': 'Success Rate',
            'solve_time_mean': 'Avg Time (s)',
            'solve_time_median': 'Median Time (s)',
            'solve_time_std': 'Time Std Dev',
            'solve_time_min': 'Min Time (s)',
            'solve_time_max': 'Max Time (s)',
            'memory_mb_mean': 'Avg Memory (MB)',
            'memory_mb_median': 'Median Memory (MB)',
            'memory_mb_max': 'Max Memory (MB)',
            'iterations_mean': 'Avg Iterations',
            'iterations_median': 'Median Iterations',
            'iterations_max': 'Max Iterations'
        }
        
        stats_df.rename(columns=column_mapping, inplace=True)
        
        # Write statistics table
        for r in dataframe_to_rows(stats_df, index=False, header=True):
            ws.append(r)
            
        # Format header
        header_row = row + 2
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
        # Percentile analysis
        row = ws.max_row + 3
        ws.cell(row=row, column=1, value="Solving Time Percentiles (seconds)")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        
        row += 2
        
        percentiles = [10, 25, 50, 75, 90, 95, 99]
        perc_data = []
        
        for algorithm in self.results_df['algorithm'].unique():
            alg_data = self.results_df[self.results_df['algorithm'] == algorithm]['solve_time']
            perc_row = {'Algorithm': algorithm}
            
            for p in percentiles:
                perc_row[f'P{p}'] = np.percentile(alg_data, p)
                
            perc_data.append(perc_row)
            
        perc_df = pd.DataFrame(perc_data).round(3)
        
        for r in dataframe_to_rows(perc_df, index=False, header=True):
            ws.append(r)
            
        # Format percentile header
        header_row = row + 2
        for cell in ws[header_row]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            
        # Correlation analysis
        row = ws.max_row + 3
        ws.cell(row=row, column=1, value="Correlation Analysis")
        ws.cell(row=row, column=1).font = Font(size=14, bold=True)
        
        row += 2
        
        # Calculate correlations
        numeric_cols = ['num_islands', 'solve_time', 'memory_mb', 'iterations']
        corr_matrix = self.results_df[numeric_cols].corr().round(3)
        
        # Write correlation matrix
        ws.cell(row=row, column=1, value="Correlation Matrix")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        # Write headers
        ws.cell(row=row, column=1, value="")
        for i, col in enumerate(corr_matrix.columns):
            ws.cell(row=row, column=i+2, value=col)
            ws.cell(row=row, column=i+2).font = Font(bold=True)
            
        # Write correlation values
        for i, (idx, row_data) in enumerate(corr_matrix.iterrows()):
            row += 1
            ws.cell(row=row, column=1, value=idx)
            ws.cell(row=row, column=1).font = Font(bold=True)
            
            for j, value in enumerate(row_data):
                ws.cell(row=row, column=j+2, value=value)
                
                # Color code correlations
                if abs(value) > 0.7:
                    ws.cell(row=row, column=j+2).fill = PatternFill(
                        start_color="FF9999", end_color="FF9999", fill_type="solid"
                    )
                elif abs(value) > 0.5:
                    ws.cell(row=row, column=j+2).fill = PatternFill(
                        start_color="FFCC99", end_color="FFCC99", fill_type="solid"
                    )
                    
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def generate_latex_report(self, filename: Optional[str] = None) -> Path:
        """
        Generate LaTeX report for academic use.
        
        Args:
            filename: Optional filename for the report
            
        Returns:
            Path to generated LaTeX file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"benchmark_report_{timestamp}.tex"
            
        filepath = self.output_dir / filename
        
        # Generate LaTeX content
        latex_content = self._generate_latex_content()
        
        # Save to file
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(latex_content)
            
        self.logger.info(f"LaTeX report saved to {filepath}")
        
        return filepath
        
    def _generate_latex_content(self) -> str:
        """Generate LaTeX report content"""
        # Calculate summary statistics
        summary_stats = self.results_df.groupby('algorithm').agg({
            'success': 'mean',
            'solve_time': ['mean', 'median'],
            'memory_mb': 'mean'
        }).round(3)
        
        latex = r"""
\documentclass[11pt]{article}
\usepackage{booktabs}
\usepackage{graphicx}
\usepackage{float}
\usepackage{amsmath}
\usepackage{hyperref}

\title{Hashiwokakero Solver Benchmark Report}
\author{Generated by Benchmark System}
\date{\today}

\begin{document}

\maketitle

\section{Executive Summary}

This report presents comprehensive benchmark results for multiple algorithms solving Hashiwokakero (Bridges) puzzles. 
The evaluation includes Integer Linear Programming (ILP), Simulated Annealing (SA), and Hybrid approaches.

\subsection{Key Findings}

\begin{itemize}
    \item Overall success rate: """ + f"{self.results_df['success'].mean() * 100:.1f}\\%" + r"""
    \item Best performing algorithm: """ + f"{summary_stats['success'].idxmax()}" + r"""
    \item Fastest average solving time: """ + f"{summary_stats['solve_time']['mean'].idxmin()}" + r"""
\end{itemize}

\section{Methodology}

\subsection{Test Configuration}
\begin{itemize}
    \item Total test cases: """ + f"{len(self.results_df)}" + r"""
    \item Algorithms tested: """ + ", ".join(self.results_df['algorithm'].unique()) + r"""
    \item Problem sizes: """ + f"{self.results_df['num_islands'].min()} to {self.results_df['num_islands'].max()} islands" + r"""
    \item Difficulty levels: """ + ", ".join(self.results_df['difficulty'].unique()) + r"""
\end{itemize}

\section{Results}

\subsection{Algorithm Performance Summary}

\begin{table}[H]
\centering
\caption{Overall Algorithm Performance}
\begin{tabular}{lrrrr}
\toprule
Algorithm & Success Rate & Avg Time (s) & Median Time (s) & Avg Memory (MB) \\
\midrule
"""
        
        # Add performance data
        for algorithm in summary_stats.index:
            success = summary_stats.loc[algorithm, ('success', 'mean')] * 100
            avg_time = summary_stats.loc[algorithm, ('solve_time', 'mean')]
            med_time = summary_stats.loc[algorithm, ('solve_time', 'median')]
            memory = summary_stats.loc[algorithm, ('memory_mb', 'mean')]
            
            latex += f"{algorithm} & {success:.1f}\\% & {avg_time:.3f} & {med_time:.3f} & {memory:.1f} \\\\\n"
            
        latex += r"""
\bottomrule
\end{tabular}
\end{table}

\subsection{Performance by Difficulty}

"""
        
        # Add difficulty analysis
        diff_analysis = pd.pivot_table(
            self.results_df,
            values='success',
            index='difficulty',
            columns='algorithm',
            aggfunc='mean'
        ) * 100
        
        latex += r"""
\begin{table}[H]
\centering
\caption{Success Rate by Difficulty (\%)}
\begin{tabular}{l""" + "r" * len(diff_analysis.columns) + r"""}
\toprule
Difficulty""" 
        
        for col in diff_analysis.columns:
            latex += f" & {col}"
            
        latex += r""" \\
\midrule
"""
        
        for idx in diff_analysis.index:
            latex += f"{idx}"
            for col in diff_analysis.columns:
                latex += f" & {diff_analysis.loc[idx, col]:.1f}"
            latex += r" \\" + "\n"
            
        latex += r"""
\bottomrule
\end{tabular}
\end{table}

\section{Conclusions}

Based on the comprehensive benchmark results:

\begin{enumerate}
    \item The """ + f"{summary_stats['success'].idxmax()}" + r""" algorithm demonstrated the highest overall success rate
    \item For time-critical applications, """ + f"{summary_stats['solve_time']['mean'].idxmin()}" + r""" provides the fastest average solving time
    \item Problem difficulty significantly impacts algorithm performance, with success rates decreasing as difficulty increases
\end{enumerate}

\end{document}
"""
        
        return latex